from openpyxl.styles import Font
from openpyxl import load_workbook
from libraries.log_manager import logger_instance, logger


class ExcelOperationManager:
    def __init__(self):
        self.workbook_cache = {}
        self.pending_operations = []

    def cache_excel_operation(self, test_step, file_path: str, actual_status: int,
                              formatted_results: str, overall_result: str,
                              formatted_fields_saved: str, test_case_manager,
                              execution_time: float) -> None:
        self.pending_operations.append({
            "tcid": test_step.get('TCID'),  # 添加TCID
            "test_step": test_step,  # 修改为test_steps
            "file_path": file_path,
            "actual_status": actual_status,
            "formatted_results": formatted_results,
            "overall_result": overall_result,
            "formatted_fields_saved": formatted_fields_saved,
            "test_case_manager": test_case_manager,
            "execution_time": execution_time
        })

    def apply_pending_operations(self) -> None:
        for operation in self.pending_operations:
            self.apply_excel_operation(**operation)

        # Save all workbooks in the cache
        for file_path, workbook in self.workbook_cache.items():
            workbook.save(file_path)

        return self.pending_operations

    def apply_excel_operation(self, tcid, test_step, file_path: str, actual_status: int,
                              formatted_results: str, overall_result: str,
                              formatted_fields_saved: str, test_case_manager,
                              execution_time: float) -> None:
        if file_path not in self.workbook_cache:
            try:
                self.workbook_cache[file_path] = load_workbook(file_path)
            except Exception as e:
                logger.error(f"Failed to load workbook: {str(e)}")
                raise
        try:
            workbook = self.workbook_cache[file_path]
            sheet = workbook.active
            actual_status_col_idx = self.get_excel_column_index(sheet, "Act Status")
            actual_result_col_idx = self.get_excel_column_index(sheet, "Act Result")
            overall_result_col_idx = self.get_excel_column_index(sheet, "Result")
            fields_saved_col_idx = self.get_excel_column_index(sheet, "Saved Fields")
            api_execution_time_col_idx = self.get_excel_column_index(sheet, "API Timing")

            row = test_case_manager.get_row_index_by_tsid(test_step["TSID"]) + 2
            sheet.cell(row=row, column=actual_status_col_idx, value=actual_status)
            sheet.cell(row=row, column=actual_result_col_idx, value=formatted_results)
            overall_cell = sheet.cell(row=row, column=overall_result_col_idx, value=overall_result)

            overall_cell.font = Font(color="006400" if overall_result == "PASS" else "8B0000")
            sheet.cell(row=row, column=fields_saved_col_idx, value=formatted_fields_saved)
            api_timing_cell = sheet.cell(row=row, column=api_execution_time_col_idx, value=f"{execution_time:.2f}s")
            if execution_time > 5:
                api_timing_cell.font = Font(color="8B0000")
            else:
                api_timing_cell.font = Font(color="006400")

        except Exception as e:
            logger.error(f"An error occurred while writing results to the Excel file: {str(e)}")
            raise

    def get_excel_column_index(self, sheet, column_name: str) -> int:
        try:
            for col in sheet.iter_cols(1, sheet.max_column):
                if col[0].value == column_name:
                    return col[0].column
        except Exception as e:
            logger.error(f"An error occurred while getting the column index for '{column_name}': {str(e)}")
            return None